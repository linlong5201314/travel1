from flask import render_template, redirect, url_for, flash, request, current_app, Blueprint, jsonify, Response
from flask_login import current_user, login_required
from datetime import datetime, timedelta
from functools import wraps
import os
from werkzeug.utils import secure_filename
import csv
from io import StringIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from sqlalchemy import or_
from utility.file_utils import save_file

# 允许上传的文件类型
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp', 'tiff', 'svg', 'heic', 'ico'}

def allowed_file(filename):
    """检查文件扩展名是否被允许"""
    # 增加更多允许的文件扩展名
    global ALLOWED_EXTENSIONS
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp', 'tiff', 'tif', 'svg', 'heic', 'ico', 
                          'jfif', 'pjpeg', 'pjp', 'avif', 'apng'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# 创建蓝图
bp = Blueprint('admin', __name__, url_prefix='/admin')

# 管理员访问权限检查
def admin_required(f):
    """装饰器：要求当前用户为管理员"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('您没有访问此页面的权限', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated_function

# 管理后台首页
@bp.route('/')
@login_required
@admin_required
def index():
    # 获取基本统计数据
    total_users = current_app.User.query.count()
    total_posts = current_app.Post.query.count()
    total_comments = current_app.Comment.query.count()
    
    # 待审核的内容
    pending_posts = current_app.Post.query.filter_by(is_approved=False, is_deleted=False).count()
    pending_comments = current_app.Comment.query.filter_by(is_approved=False, is_deleted=False).count()
    
    # 最近注册的用户
    recent_users = current_app.User.query.order_by(current_app.User.created_at.desc()).limit(5).all()
    
    # 最近的帖子
    recent_posts = current_app.Post.query.order_by(current_app.Post.created_at.desc()).limit(5).all()
    
    return render_template('admin/index.html', 
                          title='管理后台',
                          total_users=total_users,
                          total_posts=total_posts,
                          total_comments=total_comments,
                          pending_posts=pending_posts,
                          pending_comments=pending_comments,
                          recent_users=recent_users,
                          recent_posts=recent_posts,
                          now=datetime.utcnow())

# 管理用户
@bp.route('/users')
@login_required
@admin_required
def manage_users():
    page = request.args.get('page', 1, type=int)
    role = request.args.get('role', '')
    status = request.args.get('status', '')
    search_type = request.args.get('search_type', '')
    keyword = request.args.get('keyword', '')

    query = current_app.User.query.filter_by(is_deleted=False)

    # 角色筛选
    if role == 'admin':
        query = query.filter_by(role='admin')
    elif role == 'user':
        query = query.filter_by(role='user')

    # 状态筛选
    if status == 'active':
        query = query.filter_by(is_active=True)
    elif status == 'inactive':
        query = query.filter_by(is_active=False)

    # 关键词搜索
    if keyword:
        if search_type == 'email':
            query = query.filter(current_app.User.email.contains(keyword))
        else:
            query = query.filter(current_app.User.username.contains(keyword))

    users = query.order_by(current_app.User.created_at.desc()).paginate(page=page, per_page=15)
    return render_template('admin/users.html', title='用户管理', users=users, now=datetime.utcnow())

# 用户详情
@bp.route('/users/<int:user_id>')
@login_required
@admin_required
def user_detail(user_id):
    user = current_app.User.query.get_or_404(user_id)
    return render_template('admin/user_detail.html', 
                           title=f'用户详情 - {user.username}', 
                           user=user, 
                           Post=current_app.Post,
                           Comment=current_app.Comment,
                           now=datetime.utcnow())

# 管理帖子
@bp.route('/posts')
@login_required
@admin_required
def manage_posts():
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')
    search = request.args.get('search', '')
    
    query = current_app.Post.query
    
    if status == 'pending':
        query = query.filter_by(is_approved=False, is_deleted=False)
    elif status == 'approved':
        query = query.filter_by(is_approved=True, is_deleted=False)
    elif status == 'deleted':
        query = query.filter_by(is_deleted=True)
    
    # 添加搜索功能
    if search:
        query = query.filter(current_app.Post.title.contains(search))
    
    posts = query.order_by(current_app.Post.created_at.desc()).paginate(page=page, per_page=15)
    return render_template('admin/posts.html', title='文章管理', posts=posts, status=status, now=datetime.utcnow())

# 查看帖子详情
@bp.route('/posts/<int:post_id>')
@login_required
@admin_required
def post_detail(post_id):
    post = current_app.Post.query.get_or_404(post_id)
    return render_template('admin/post_detail.html', title=f'帖子详情 - {post.title}', post=post, Comment=current_app.Comment, request=request, now=datetime.utcnow())

# 帖子前台预览
@bp.route('/posts/<int:post_id>/preview')
@login_required
@admin_required
def post_preview(post_id):
    post = current_app.Post.query.get_or_404(post_id)
    
    # 获取已审核的评论
    comments = current_app.Comment.query.filter_by(
        post_id=post_id, 
        parent_id=None, 
        is_deleted=False,
        is_approved=True
    ).order_by(current_app.Comment.created_at.desc()).all()
    
    # 管理员可以看到所有评论，包括未审核的
    pending_comments = current_app.Comment.query.filter_by(
        post_id=post_id, 
        parent_id=None, 
        is_deleted=False,
        is_approved=False
    ).order_by(current_app.Comment.created_at.desc()).all()
    
    all_comments = pending_comments + comments
    
    # 获取相关故事（同一位置或同一作者的其他故事）
    related_posts = current_app.Post.query.filter(
        current_app.Post.id != post_id,
        current_app.Post.is_approved == True,
        current_app.Post.is_deleted == False,
        or_(
            current_app.Post.user_id == post.user_id,
            current_app.Post.location == post.location
        )
    ).order_by(current_app.Post.created_at.desc()).limit(3).all()
    
    return render_template('admin/post_preview.html', 
                          title=f'前台预览 - {post.title}', 
                          post=post, 
                          comments=all_comments,
                          related_posts=related_posts,
                          Comment=current_app.Comment,
                          now=datetime.utcnow())

# 管理评论
@bp.route('/comments')
@login_required
@admin_required
def manage_comments():
    page = request.args.get('page', 1, type=int)
    status = request.args.get('status', '')
    # 获取搜索关键字
    search = request.args.get('search', '').strip()
    
    query = current_app.Comment.query
    
    if status == 'pending':
        query = query.filter_by(is_approved=False, is_deleted=False)
    elif status == 'approved':
        query = query.filter_by(is_approved=True, is_deleted=False)
    elif status == 'deleted':
        query = query.filter_by(is_deleted=True)
    
    # 根据关键字搜索评论内容（支持模糊匹配）
    if search:
        query = query.filter(current_app.Comment.content.contains(search))
    
    comments = query.order_by(current_app.Comment.created_at.desc()).paginate(page=page, per_page=15)
    # 无需显式传递 search，因为模板可通过 request.args 获取，但为了模板使用方便，这里也一并传入
    return render_template('admin/comments.html', title='评论管理', comments=comments, status=status, now=datetime.utcnow())

# 审核帖子
@bp.route('/approve_post/<int:post_id>', methods=['POST'])
@login_required
@admin_required
def approve_post(post_id):
    post = current_app.Post.query.get_or_404(post_id)
    post.is_approved = True
    current_app.db.session.commit()
    flash('文章已通过审核', 'success')
    return redirect(url_for('admin.manage_posts', status='pending'))

# 拒绝帖子
@bp.route('/reject_post/<int:post_id>', methods=['POST'])
@login_required
@admin_required
def reject_post(post_id):
    post = current_app.Post.query.get_or_404(post_id)
    post.is_deleted = True
    current_app.db.session.commit()
    flash('文章已被拒绝', 'success')
    return redirect(url_for('admin.manage_posts', status='pending'))

# 审核评论
@bp.route('/approve_comment/<int:comment_id>', methods=['POST'])
@login_required
@admin_required
def approve_comment(comment_id):
    comment = current_app.Comment.query.get_or_404(comment_id)
    comment.is_approved = True
    current_app.db.session.commit()
    flash('评论已通过审核', 'success')
    return redirect(url_for('admin.manage_comments', status='pending'))

# 拒绝评论
@bp.route('/reject_comment/<int:comment_id>', methods=['POST'])
@login_required
@admin_required
def reject_comment(comment_id):
    comment = current_app.Comment.query.get_or_404(comment_id)
    comment.is_deleted = True
    current_app.db.session.commit()
    flash('评论已被拒绝', 'success')
    return redirect(url_for('admin.manage_comments', status='pending'))

# 禁用/启用用户
@bp.route('/toggle_user_status/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def toggle_user_status(user_id):
    user = current_app.User.query.get_or_404(user_id)
    
    # 不能禁用自己
    if user.id == current_user.id:
        flash('不能禁用自己的账号', 'danger')
        return redirect(url_for('admin.manage_users'))
    
    user.is_active = not user.is_active
    current_app.db.session.commit()
    
    status_msg = '启用' if user.is_active else '禁用'
    flash(f'用户 {user.username} 已{status_msg}', 'success')
    return redirect(url_for('admin.manage_users'))

# 删除帖子
@bp.route('/delete_post/<int:post_id>', methods=['POST'])
@login_required
@admin_required
def delete_post(post_id):
    post = current_app.Post.query.get_or_404(post_id)
    post.is_deleted = True
    current_app.db.session.commit()
    flash('帖子已成功删除', 'success')
    
    # 如果有return_to参数，则返回到指定URL
    return_to = request.args.get('return_to')
    if return_to:
        return redirect(return_to)
    return redirect(url_for('admin.manage_posts'))

# 编辑帖子
@bp.route('/edit_post/<int:post_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_post(post_id):
    post = current_app.Post.query.get_or_404(post_id)
    if request.method == 'POST':
        post.title = request.form.get('title')
        post.content = request.form.get('content')
        post.location = request.form.get('location')
        post.is_approved = True if request.form.get('is_approved') == 'on' else False
        
        # 处理新增字段
        post.category = request.form.get('category')
        post.destination_type = request.form.get('destination_type')
        
        # 处理评分字段 (如果提交的表单中有评分)
        rating = request.form.get('rating')
        if rating:
            try:
                post.rating = float(rating)
            except ValueError:
                # 如果评分不是有效的浮点数，则默认为0
                post.rating = 0.0
        
        # 处理封面图片上传
        featured_file = request.files.get('featured_image')
        if featured_file and featured_file.filename:
            print(f"DEBUG: 尝试上传封面图片: {featured_file.filename}")
            # 直接打印文件对象信息
            print(f"DEBUG: 文件信息: {featured_file}, MIME类型: {featured_file.content_type if hasattr(featured_file, 'content_type') else '未知'}")
            
            # 检查文件类型并允许任何图像类型
            file_ext = featured_file.filename.rsplit('.', 1)[1].lower() if '.' in featured_file.filename else ''
            print(f"DEBUG: 文件扩展名: {file_ext}")
            
            unique_filename = save_file(featured_file, 'uploads/posts')
            if unique_filename:
                print(f"DEBUG: 封面图片上传成功: {unique_filename}")
                post.featured_image = unique_filename
            else:
                print(f"DEBUG: 封面图片上传失败，文件类型可能不支持或保存出错")
                flash('封面图片上传失败，文件类型不支持', 'danger')
        
        # 处理多张附加图片
        images = request.files.getlist('images')
        for img in images:
            unique_filename = save_file(img, 'uploads/posts')
            if unique_filename:
                post_image = current_app.PostImage(filename=unique_filename)
                post.images.append(post_image)
        
        # 如果尚未设置封面，则自动使用第一张附图
        if not post.featured_image and post.images:
            post.featured_image = post.images[0].filename
        
        current_app.db.session.commit()
        flash('帖子已更新', 'success')
        return redirect(url_for('admin.post_detail', post_id=post.id))
    return render_template('admin/edit_post.html', title='编辑帖子', post=post, now=datetime.utcnow())

# 删除评论
@bp.route('/delete_comment/<int:comment_id>', methods=['POST'])
@login_required
@admin_required
def delete_comment(comment_id):
    comment = current_app.Comment.query.get_or_404(comment_id)
    comment.is_deleted = True
    current_app.db.session.commit()
    flash('评论已成功删除', 'success')
    
    # 如果有return_to参数，则返回到指定URL
    return_to = request.form.get('return_to') or request.args.get('return_to')
    if return_to:
        return redirect(return_to)
    
    # 如果评论所属的帖子ID可用，则返回到帖子详情页
    if comment.post_id:
        return redirect(url_for('admin.post_detail', post_id=comment.post_id))
    
    # 默认返回到评论管理页面
    return redirect(url_for('admin.manage_comments'))

# 编辑评论
@bp.route('/edit_comment/<int:comment_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_comment(comment_id):
    comment = current_app.Comment.query.get_or_404(comment_id)
    if request.method == 'POST':
        comment.content = request.form.get('content')
        comment.is_approved = True if request.form.get('is_approved') == 'on' else False
        current_app.db.session.commit()
        flash('评论已更新', 'success')
        return redirect(url_for('admin.manage_comments'))
    return render_template('admin/edit_comment.html', title='编辑评论', comment=comment, now=datetime.utcnow())

# 添加用户
@bp.route('/add_user', methods=['POST'])
@login_required
@admin_required
def add_user():
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    role = request.form.get('role', 'user')
    
    # 检查用户名是否已存在
    existing_user = current_app.User.query.filter_by(username=username, is_deleted=False).first()
    if existing_user:
        flash('用户名已存在', 'danger')
        return redirect(url_for('admin.manage_users'))
    
    # 检查邮箱是否已存在
    existing_email = current_app.User.query.filter_by(email=email, is_deleted=False).first()
    if existing_email:
        flash('邮箱已被注册', 'danger')
        return redirect(url_for('admin.manage_users'))
    
    # 创建新用户
    new_user = current_app.User(username=username, email=email, password=password, role=role)
    current_app.db.session.add(new_user)
    current_app.db.session.commit()
    
    flash('用户添加成功', 'success')
    return redirect(url_for('admin.manage_users'))

# 删除用户
@bp.route('/delete_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    user = current_app.User.query.get_or_404(user_id)
    
    # 不能删除自己
    if user.id == current_user.id:
        flash('不能删除自己的账号', 'danger')
        return redirect(url_for('admin.manage_users'))
    
    # 标记用户为已删除
    user.is_deleted = True
    current_app.db.session.commit()
    
    flash(f'用户 {user.username} 已成功删除', 'success')
    return redirect(url_for('admin.manage_users'))

# 编辑用户
@bp.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    user = current_app.User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role', 'user')
        is_active = True if request.form.get('is_active') else False
        
        # 检查用户名是否已存在（排除当前用户）
        existing_user = current_app.User.query.filter(
            current_app.User.username == username, 
            current_app.User.id != user_id,
            current_app.User.is_deleted == False
        ).first()
        
        if existing_user:
            flash('用户名已存在', 'danger')
            return render_template('admin/edit_user.html', title='编辑用户', user=user, now=datetime.utcnow())
        
        # 检查邮箱是否已存在（排除当前用户）
        existing_email = current_app.User.query.filter(
            current_app.User.email == email, 
            current_app.User.id != user_id,
            current_app.User.is_deleted == False
        ).first()
        
        if existing_email:
            flash('邮箱已被注册', 'danger')
            return render_template('admin/edit_user.html', title='编辑用户', user=user, now=datetime.utcnow())
        
        # 处理头像上传
        avatar_file = request.files.get('avatar')
        if avatar_file and avatar_file.filename:
            unique_filename = save_file(avatar_file, 'uploads/profiles')
            if unique_filename:
                # 删除旧头像
                if user.avatar_url:
                    from utility.file_utils import delete_file
                    delete_file(user.avatar_url, 'uploads/profiles')
                user.avatar_url = unique_filename
        
        # 更新用户信息
        user.username = username
        user.email = email
        if password:  # 如果提供了新密码
            user.set_password(password)
        user.role = role
        user.is_active = is_active
        
        current_app.db.session.commit()
        
        flash('用户信息已更新', 'success')
        return redirect(url_for('admin.user_detail', user_id=user.id))
    
    return render_template('admin/edit_user.html', title='编辑用户', user=user, now=datetime.utcnow())

@bp.route('/admin_analytics')
@login_required
@admin_required
@cache.cached(timeout=300, key_prefix='admin_analytics')
def admin_analytics():
    """管理员数据可视化页面"""
    try:
        # 获取基础统计数据
        user_count = current_app.User.query.count()
        post_count = current_app.Post.query.count()
        comment_count = current_app.Comment.query.count()
        bookmark_count = current_app.Bookmark.query.count()
        # 计算总浏览量
        total_views = current_app.db.session.query(current_app.db.func.sum(current_app.Post.views)).scalar() or 0
        # 获取活跃用户
        active_users = current_app.User.query.filter_by(is_active=True).count()
        # 计算过去7天的新增用户
        week_ago = datetime.utcnow() - timedelta(days=7)
        new_users_week = current_app.User.query.filter(current_app.User.created_at >= week_ago).count()
        # 计算过去30天的新增用户
        month_ago = datetime.utcnow() - timedelta(days=30)
        new_users_month = current_app.User.query.filter(current_app.User.created_at >= month_ago).count()
        # 获取用户角色分布
        admin_count = current_app.User.query.filter_by(role='admin').count()
        user_role_data = {
            'labels': ['管理员', '普通用户'],
            'data': [admin_count, user_count - admin_count]
        }
        # 获取过去6个月的用户注册趋势和文章发布趋势
        months_data = []
        post_months_data = []
        months_labels = []
        for i in range(5, -1, -1):
            month_start = datetime.utcnow().replace(day=1) - timedelta(days=i*30)
            month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
            # 一次循环中获取两种数据，减少循环次数
            user_month_count = current_app.User.query.filter(current_app.User.created_at >= month_start, current_app.User.created_at < month_end).count()
            post_month_count = current_app.Post.query.filter(current_app.Post.created_at >= month_start, current_app.Post.created_at < month_end).count()
            
            months_data.append(user_month_count)
            post_months_data.append(post_month_count)
            months_labels.append(month_start.strftime('%Y-%m'))
        
        post_trend_data = {
            'labels': months_labels,
            'data': post_months_data
        }
        
        # 获取文章浏览量最高的前5篇
        top_posts = current_app.Post.query.order_by(current_app.Post.views.desc()).limit(5).all()
        
        # 获取评论最多的前5篇文章
        most_commented = current_app.db.session.query(
            current_app.Post, current_app.db.func.count(current_app.Comment.id).label('comment_count')
        ).join(current_app.Comment).group_by(current_app.Post.id).order_by(current_app.db.desc('comment_count')).limit(5).all()
        
        # 获取地区分布
        locations = current_app.db.session.query(
            current_app.Post.location, current_app.db.func.count(current_app.Post.id).label('location_count')
        ).filter(current_app.Post.location != None).filter(current_app.Post.location != '').group_by(current_app.Post.location).order_by(current_app.db.desc('location_count')).limit(10).all()
        
        location_data = {
            'labels': [loc[0] for loc in locations if loc[0]],
            'data': [loc[1] for loc in locations if loc[0]]
        }
        
        # 创建前端计算需要用到的比率数据
        # 避免除以零错误
        new_users_month_ratio = 0
        if user_count > 0:
            new_users_month_ratio = (new_users_month / user_count) * 100
        avg_comments_per_post = 0
        if post_count > 0:
            avg_comments_per_post = comment_count / post_count
        
        # 文章分类分布统计
        category_stats = current_app.db.session.query(
            current_app.Post.category, current_app.db.func.count(current_app.Post.id)
        ).group_by(current_app.Post.category).all()
        category_data = [
            {'name': t[0] or '未知', 'value': t[1]} for t in category_stats if t[0]
        ]
        
        # 用户活跃度分布统计
        user_activity_stats = [
            {'name': '每日活跃', 'value': current_app.User.query.filter(current_app.User.activity_score >= 80).count()},
            {'name': '每周活跃', 'value': current_app.User.query.filter(current_app.User.activity_score >= 60, current_app.User.activity_score < 80).count()},
            {'name': '每月活跃', 'value': current_app.User.query.filter(current_app.User.activity_score >= 40, current_app.User.activity_score < 60).count()},
            {'name': '不定期活跃', 'value': current_app.User.query.filter(current_app.User.activity_score >= 20, current_app.User.activity_score < 40).count()},
            {'name': '长期不活跃', 'value': current_app.User.query.filter(current_app.User.activity_score < 20).count()}
        ]
        
        # 文章评分分布统计
        rating_stats = [
            {'name': '5星', 'value': current_app.Post.query.filter(current_app.Post.rating >= 4.5).count()},
            {'name': '4星', 'value': current_app.Post.query.filter(current_app.Post.rating >= 3.5, current_app.Post.rating < 4.5).count()},
            {'name': '3星', 'value': current_app.Post.query.filter(current_app.Post.rating >= 2.5, current_app.Post.rating < 3.5).count()},
            {'name': '2星', 'value': current_app.Post.query.filter(current_app.Post.rating >= 1.5, current_app.Post.rating < 2.5).count()},
            {'name': '1星', 'value': current_app.Post.query.filter(current_app.Post.rating < 1.5).count()}
        ]
        
        # 目的地类型分布统计
        destination_type_stats = current_app.db.session.query(
            current_app.Post.destination_type, current_app.db.func.count(current_app.Post.id)
        ).group_by(current_app.Post.destination_type).all()
        
        destination_type_data = [
            {'name': t[0] or '未知', 'value': t[1]} for t in destination_type_stats if t[0]
        ]
        
        # 渲染模板
        return render_template(
            'admin/analytics.html',
            user_count=user_count,
            post_count=post_count,
            comment_count=comment_count,
            bookmark_count=bookmark_count,
            total_views=total_views,
            active_users=active_users,
            new_users_week=new_users_week,
            new_users_month=new_users_month,
            new_users_month_ratio=new_users_month_ratio,
            avg_comments_per_post=avg_comments_per_post,
            user_role_data=user_role_data,
            months_labels=months_labels,
            months_data=months_data,
            top_posts=top_posts,
            most_commented=most_commented,
            location_data=location_data,
            post_trend_data=post_trend_data,
            destination_type_data=destination_type_data,
            category_data=category_data,
            user_activity_stats=user_activity_stats,
            rating_stats=rating_stats,
            config=current_app.config
        )
    except Exception as e:
        flash(f'加载数据分析页面时出现错误: {str(e)}', 'danger')
        return redirect(url_for('admin.index'))

@bp.route('/export/top_posts')
@login_required
@admin_required
def export_top_posts():
    top_posts = current_app.Post.query.order_by(current_app.Post.views.desc()).limit(5).all()
    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(['文章ID', '文章标题', '发布日期', '作者', '阅读量', '点赞数'])
    for post in top_posts:
        cw.writerow([
            post.id,
            post.title,
            post.created_at.strftime('%Y-%m-%d'),
            post.author.username if post.author else '未知',
            post.views,
            post.like_count()
        ])
    output = si.getvalue().encode('utf-8-sig')
    return Response(output, mimetype='text/csv', headers={
        'Content-Disposition': 'attachment; filename=top_posts.csv'
    })

@bp.route('/export/location_stats')
@login_required
@admin_required
def export_location_stats():
    locations = current_app.db.session.query(
        current_app.Post.location, current_app.db.func.count(current_app.Post.id).label('location_count')
    ).filter(current_app.Post.location != None).filter(current_app.Post.location != '').group_by(current_app.Post.location).order_by(current_app.db.desc('location_count')).limit(10).all()
    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(['目的地', '文章数量'])
    for loc in locations:
        cw.writerow([loc[0], loc[1]])
    output = si.getvalue().encode('utf-8-sig')
    return Response(output, mimetype='text/csv', headers={
        'Content-Disposition': 'attachment; filename=location_stats.csv'
    }) 

@bp.route('/export/post/<int:post_id>')
@login_required
@admin_required
def export_post_detail(post_id):
    post = current_app.Post.query.get_or_404(post_id)
    comments = post.comments.filter_by(is_deleted=False).order_by(current_app.Comment.created_at.asc()).all()
    si = StringIO()
    cw = csv.writer(si)
    # 帖子基本信息
    cw.writerow(['帖子ID', '标题', '作者', '发布时间', '内容', '浏览量', '点赞数', '评论数'])
    cw.writerow([
        post.id,
        post.title,
        post.author.username if post.author else '未知',
        post.created_at.strftime('%Y-%m-%d %H:%M'),
        post.content.replace('\n', ' ').replace('\r', ' '),
        post.views,
        post.like_count() if hasattr(post, 'like_count') else '',
        post.comments.filter_by(is_deleted=False).count()
    ])
    # 评论信息
    cw.writerow([])
    cw.writerow(['评论ID', '评论用户', '评论时间', '评论内容'])
    for c in comments:
        cw.writerow([
            c.id,
            c.author.username if c.author else '未知',
            c.created_at.strftime('%Y-%m-%d %H:%M'),
            c.content.replace('\n', ' ').replace('\r', ' ')
        ])
    output = si.getvalue().encode('utf-8-sig')
    return Response(output, mimetype='text/csv', headers={
        'Content-Disposition': f'attachment; filename=post_{post.id}_detail.csv'
    })

@bp.route('/export/post_excel/<int:post_id>')
@login_required
@admin_required
def export_post_detail_excel(post_id):
    post = current_app.Post.query.get_or_404(post_id)
    comments = post.comments.filter_by(is_deleted=False).order_by(current_app.Comment.created_at.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "帖子详情"

    # 帖子基本信息
    ws.append(['帖子ID', '标题', '作者', '发布时间', '内容', '浏览量', '点赞数', '评论数'])
    ws.append([
        post.id,
        post.title,
        post.author.username if post.author else '未知',
        post.created_at.strftime('%Y-%m-%d %H:%M'),
        post.content.replace('\n', ' ').replace('\r', ' '),
        post.views,
        post.like_count() if hasattr(post, 'like_count') else '',
        post.comments.filter_by(is_deleted=False).count()
    ])

    # 空行
    ws.append([])
    ws.append(['评论ID', '评论用户', '评论时间', '评论内容'])
    for c in comments:
        ws.append([
            c.id,
            c.author.username if c.author else '未知',
            c.created_at.strftime('%Y-%m-%d %H:%M'),
            c.content.replace('\n', ' ').replace('\r', ' ')
        ])

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename=post_{post.id}_detail.xlsx'
        }
    ) 